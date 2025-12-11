#!/usr/bin/env python3
import torch
import torch.nn as nn
import torch.nn.functional as F
from torchinfo import summary
import numpy as np
import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from torch.utils.data import Dataset, DataLoader
import plotly.graph_objects as go
from cPlot import debug_plot
from scipy.signal import find_peaks as sp_find_peaks
from tqdm.auto import tqdm

# ——————————————————————————————————————————————————————————————
# 0) Configuration dictionary
# ——————————————————————————————————————————————————————————————
def Generate_config():
    config = {
        # array & data
        'N':               5,           # sensors
        'T':               8,           # snapshots per sample
        'L':            1801,           # grid size
        'K':               2,           # sources per sample (overridden if fixed GT)
        'SNR_dB_range':  (30.0, 30.0),  # desired SNR range in dB 

        # fixed GT (or None for random)
        'gt_angles_deg':   None,
        'th_md_deg':        10.0,

        # debugging
        'debug_plot': False, # True , False
        'debug_loss': False,

        # training / unfolding
        'num_train':     3000,    # 1000
        'num_test':      200,     # 200
        'num_epochs':    100,     # 100
        'batch_size':    128,      # 64

        'param_predictor_module': 'ComplexConv',  # FullyConnected , MLP , Transformer , Mixer , ComplexConv

        # 'num_train':     16,    # 1000
        # 'num_test':      8,     # 200
        # 'num_epochs':    10,    # 100
        # 'batch_size':    4,     # 128

        'num_layers':    17,     # unfolding depth
        'learning_rate': 1e-1,
        'weight_decay':  1e-2,

        # algorithm selection
        'algorithm':    'MFOCUSS',       # 'IAA' or 'MFOCUSS'
        'mode':         'unfolded',   # 'unfolded' or 'classic'
        'param_predictor': 'mlp',      # 'mlp', 'transformer', 'mixer', 'complexconv'
        'classic_iters': 100,          # iterations for classic

        # --- DNN options ---
        'use_residual':    False,     # (1) u = p + γ*(v-p)
        'use_layernorm':   False,     # (2)
        'use_attention':   False,     # (3)
        'use_gating':      False,     # (4)
        'use_shrinkage':   False,     # (5)
        'constrain_gamma': False,     # γ = sigmoid(raw)
        
        'delta_init':       1.0,     # initial raw δ

        # IAA
        'use_ridge':       False,     # +ridge on R
        'ridge_reg':       1e-3,     # δ
        
        'lmbda_init':       0.1,     # IAA shrinkage λ init

        # --- Classic M-FOCUSS params ---
        'max_est_peaks':       2,
        'p_decay_rate':        0.1,    # p ← p_init * exp(-p_decay_rate * k)
        'lambda_growth_rate':  0.1,    # λ growth exponent

        'min_df_dist_combine_deg': 2.0,    # merge any two DOAs closer than 5°
        'max_num_of_est_doas':     7,

        'mfocuss_p_init':      1e-3,   # 0.1,
        'mfocuss_lambda_init': 1e-2,   # 0.1,      # λ₀
        
        'p_min_th':            1e-3,   # stop if p ≤ this
        'lambda_max_th':       1e-2,   # stop-growth threshold

        'epsilon':             1e-3,   # frobenius‐norm convergence tol.
        's_norm_factor':       1.0,    # input normalization scale
        'prune_gamma_th_dB':   -20.0,  # gamma pruning threshold in dB
        'gap_from_peak_dB':    20.0,

        # energy normalization
        'preserve_energy': False,

        # M-FOCUSS “unfolded” hyper‐limits (if you still use it)
        # 'mfocuss_p_min':       0.001,
        # 'mfocuss_p_max':       0.99,
        
        # 'mfocuss_lambda_min':  0.0,
        # 'mfocuss_lambda_max':  1e-2,

        'constrain_p_lam': True,

        'mfocuss_p_min':  0.05,   # keep away from 0
        'mfocuss_p_max':  0.95,    # strictly < 2 to avoid negative exponent issues

        'mfocuss_lambda_min': 1e-7,
        'mfocuss_lambda_max': 1e-2,

        # CS normalization parameters
        's_norm_factor': {
            'nfMode': 'All',  # Can be "Legacy", "Elements", or "All"
            'nfRMS': 15.0,     # RMS normalization factor
            'nfLegacyRMS': 100.0  # Legacy RMS normalization factor
        },
    }

# ——————————————————————————————————————————————————————————————
# 0b) Global angle grid
# ——————————————————————————————————————————————————————————————
    theta_deg = np.linspace(-90.0, 90.0, config['L'])
    theta_rad = theta_deg * np.pi/180.0
    config['theta_grid_deg'] = theta_deg
    config['theta_grid_rad'] = theta_rad
    return config

eps = 1e-12 # np.finfo(float).eps
config = Generate_config()

# ——————————————————————————————————————————————————————————————
# 1) Ground‐truth angle normalization helper
# ——————————————————————————————————————————————————————————————
def Update_gt_angles(cfg):
    gt = cfg.get('gt_angles_deg')
    if gt is not None:
        if isinstance(gt, (float, int)):
            gt = [float(gt)]
        elif isinstance(gt, np.ndarray):
            gt = gt.tolist()
        cfg['gt_angles_deg'] = gt
        cfg['K'] = len(gt)

# ——————————————————————————————————————————————————————————————
# 2) logger for scalar params × layers
# ——————————————————————————————————————————————————————————————

def Limit_decimal_digits(ws):

     # --- Format every float to at most 3 decimal places ---
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = '0.###'

def Conditional_formatting(ws, n_samples, n_sub):
    start_col = 2
    end_col   = start_col + n_samples - 1
    start_row = 2
    end_row   = start_row + n_sub - 1
    cell_range = (
        f"{get_column_letter(start_col)}{start_row}:"
        f"{get_column_letter(end_col)}{end_row}"
    )
    rule = ColorScaleRule(
        start_type="min",      start_color="63BE7B",  # green
        mid_type="percentile", mid_value=50,           mid_color="FFEB84",  # yellow
        end_type="max",        end_color="F8696B"      # red
    )
    ws.conditional_formatting.add(cell_range, rule)

def Auto_fit_column_widths(ws):
    for col_cells in ws.columns:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = max_len + 2

def save_history(history: dict,
                 tags: list,
                 log_dir: str,
                 filename: str,
                 mode: str = "Train"):
    """
    history: dict[name] -> list of lists (each inner list has length = n_sub)
    tags:    list of strings, one per “row” (either epoch_batch tags or sample tags)
    mode:    "Train" or "Test" to choose row‐index naming
    """
    os.makedirs(log_dir, exist_ok=True)
    path = os.path.join(log_dir, filename)
    print(f"→ Writing {mode} history to {path}")

    with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
        for name, hist in history.items():
            # hist: List[List], shape = (n_samples, n_sub)
            n_samples = len(hist)
            n_sub     = len(hist[0]) if n_samples>0 else 0

            # build DataFrame; pandas will pad ragged entries with NaN
            df = pd.DataFrame(
                hist,
                index=tags,
                columns=[f"{name}_{j}" for j in range(n_sub)]
            )
            df.index.name = "sample" if mode=="Test" else "epoch_batch"

            # transpose so rows = sub‐items, cols = samples
            df_t = df.T
            df_t.to_excel(writer, sheet_name=name)

            ws = writer.sheets[name]

            # 1) Auto‐fit column widths
            Auto_fit_column_widths(ws)

            # 2) Conditional formatting over the data block B2:<col><row>
            Conditional_formatting(ws, n_samples, n_sub)

            # 3) Limit all floats to 3 decimal digits
            Limit_decimal_digits(ws)

    print("→ Done.")

def Update_history(history, results, net):
    epoch_vals = {}
    for layer in net.layers:
        for name, param in layer.named_parameters(recurse=False):
            v = param.detach().cpu().item()
            epoch_vals.setdefault(name, []).append(v)
    for name, vals in epoch_vals.items():
        history.setdefault(name, []).append(vals)

    for name, arr in results.items():
        a = arr.detach().cpu().numpy() if hasattr(arr, "detach") else np.asarray(arr)
        history.setdefault(name, []).append(a.flatten().tolist())

from torchinfo import summary

def show_model_summary(model: nn.Module, input_shape: tuple, model_name: str = "Param Predictor", is_complex=False):
    print(f"\n{'='*60}")
    print(f"📦 Model Summary: {model_name}")
    print(f"{'='*60}")

    if is_complex:
        dummy_input = torch.randn(input_shape, dtype=torch.complex64)
    else:
        dummy_input = torch.randn(input_shape)

    # ✅ Only add 1 batch dimension
    if dummy_input.dim() == 3:
        dummy_input = dummy_input.unsqueeze(0)  # (1, C, H, W)
    elif dummy_input.dim() == 2:
        dummy_input = dummy_input.unsqueeze(0)  # (1, input_dim)

    summary(model,
            input_data=dummy_input,
            col_names=["input_size", "output_size", "num_params"],
            depth=4)

    print(f"{'='*60}\n")

# def show_model_summary(model: nn.Module, input_shape: tuple, model_name: str = "Param Predictor", is_complex=False):
#     print(f"\n{'='*60}")
#     print(f"📦 Model Summary: {model_name}")
#     print(f"{'='*60}")

#     # Create dummy input with correct dtype
#     if is_complex:
#         dummy_input = torch.randn(input_shape, dtype=torch.complex64)
#     else:
#         dummy_input = torch.randn(input_shape)

#     # Add batch dimension
#     dummy_input = dummy_input.unsqueeze(0)

#     summary(model,
#             input_data=dummy_input,
#             col_names=["input_size", "output_size", "num_params"],
#             depth=4)

#     print(f"{'='*60}\n")


# def show_model_summary(model: nn.Module, input_shape: tuple, model_name: str = "Param Predictor"):
#     """
#     Display the architecture summary of a learning model.

#     Parameters:
#     - model (nn.Module): the instantiated PyTorch model (MLP, Transformer, etc.)
#     - input_shape (tuple): the shape of the model input (excluding batch dimension)
#     - model_name (str): optional name to display
#     """
#     print(f"\n{'='*60}")
#     print(f"📦 Model Summary: {model_name}")
#     print(f"{'='*60}")
#     summary(model, input_size=(1, *input_shape), col_names=["input_size", "output_size", "num_params"], depth=4)
#     print(f"{'='*60}\n")


# ——————————————————————————————————————————————————————————————
# 3) Debug plots
# ——————————————————————————————————————————————————————————————
def _build_plot_title(*, results: dict = None, epoch: int = None, loss: float = None):
    """
    Assemble a list of strings from epoch, loss, and any
    key→value pairs in results, then join with “ | ”.
    """
    parts = []
    if epoch is not None:
        parts.append(f"Epoch {epoch}")
    if loss is not None:
        # parts.append(f"Loss={pow2db(loss):.4e} [dB]")
        parts.append(f"Loss={pow2db(loss):.2} [dB]")
    if results is not None:
        for k, v in results.items():
            # pretty‐print small arrays
            if isinstance(v, np.ndarray):
                lst = [float(f"{x:.1f}") for x in v.tolist()]
                parts.append(f"{k}={lst}")
            elif isinstance(v, float):
                parts.append(f"{k}={v:.4f}")
            else:
                parts.append(f"{k}={v}")
    return " | ".join(parts)

def pow2db(lin):
    dB = 10 * np.log10(lin.detach().cpu().numpy() + eps)
    return dB

def mag2db(lin):
    
    dB = 20 * np.log10(lin.detach().cpu().numpy() + eps)
    return dB

def Plot_gamma_hat_vs_true(gamma_true, gamma_hat, results=None, epoch=None, loss=None):

    # protect against results being None
    results = results or {}

    # compute dB spectra
    g_true_dB = pow2db(gamma_true + eps)
    g_hat_dB  = pow2db(gamma_hat + eps)
    eps_dB = 10*np.log10(eps)

    # clamp anything below –40 dB down to –120 dB
    g_hat_dB = np.where(g_hat_dB < -40.0, eps_dB, g_hat_dB)

    # stack into 2×L array for two curves
    arr = np.vstack([g_true_dB, g_hat_dB])

    # drop the two keys (no error if missing)
    plot_results = results.copy()
    plot_results.pop('Est_DOA_deg',   None)
    plot_results.pop('Peaks_inds',     None)

    # build title from results / epoch / loss
    title = _build_plot_title(results=plot_results, epoch=epoch, loss=loss) \
            or "γ_true vs γ_hat"

    # pull out and clean up vertical‐line positions
    gt_lines  = results.get('GT_DOA_deg', [])
    gt_lines  = np.atleast_1d(gt_lines).tolist()

    est_lines = results.get('Est_DOA_deg', [])
    est_lines = np.atleast_1d(est_lines).flatten()
    est_lines = [float(x) for x in est_lines if not np.isnan(x)]

    # tell debug_plot we're in curve‐mode, give it names for the two traces
    axes = {
        'mode'        : 'curves',
        'title'       : title,
        'x'           : config['theta_grid_deg'],
        'names'       : ['γ_true (dB)', 'γ_hat (dB)'],
        'x_label'     : "Angle", 'x_unit': "°",
        'y_label'     : "Power", 'y_unit': "dB",
        'vlines_1'    : gt_lines,
        'color_1'     : 'green',
        'vlines_2'    : est_lines,
        'color_2'     : 'red',
    }

    debug_plot(arr, axes)

def Show_gamma(gamma_progress, results=None, *, epoch=None, loss=None):
    # layers×L heatmap
    mat = torch.stack(gamma_progress, dim=0).detach().cpu().numpy()    

    # drop the two keys (no error if missing)
    plot_results = results.copy()
    plot_results.pop('Est_DOA_deg',   None)
    plot_results.pop('Peaks_inds',     None)

    title = _build_plot_title(results=plot_results, epoch=epoch, loss=loss)
    title = title or "Layer‐by‐Layer γ Progress"

    # pull out your two sets of v‐lines
    gt_lines  = results.get('GT_DOA_deg', [])
    est_lines = results.get('Est_DOA_deg', [])

    # drop any NaNs from the estimates
    est_lines = np.atleast_1d(est_lines).flatten()
    est_lines = [float(x) for x in est_lines if not np.isnan(x)]

    axes = {
        'mode'        : 'heatmap',
        'title'       : title,
        'x'           : config['theta_grid_deg'],
        'y'           : list(range(1, len(gamma_progress)+1)),
        'x_label'     : "Angle", 'x_unit': "°",
        'y_label'     : "Layer", 'y_unit': "",
        'vlines_1'    : gt_lines,
        'color_1'     : 'green',
        'vlines_2'    : est_lines,
        'color_2'     : 'red',
    }
    debug_plot(mat, axes)

def Plot_gamma(gamma):
    axes = {
        'x':       config['theta_grid_deg'],
        'title':   "M-FOCUSS Gamma",
        'x_label': "Angle",   'x_unit': "°",
        'y_label': "Power",   'y_unit': "",
        'vlines_1':  config.get('gt_angles_deg'),
        'color_1': 'green'
    }
    debug_plot(gamma, axes)

def Show_power_spectrum(gamma_progress, results, g_true, g_hat):

    Show_gamma(
        gamma_progress,
        results=results,
    )
    Plot_gamma_hat_vs_true(
        g_true,
        g_hat,
        results=results,
    )

# ——————————————————————————————————————————————————————————————
# 4) Core init for IAA & M-FOCUSS
# ——————————————————————————————————————————————————————————————
def IAA_initial_power(A, X):
    N, T = X.shape
    _, L = A.shape
    XXh = X @ X.conj().T

    gamma_init = torch.zeros(L, device=A.device)
    for l in range(L):
        a = A[:, l:l+1]
        num = torch.real((a.conj().T @ XXh @ a).squeeze())
        den = torch.real((a.conj().T @ a).squeeze())
        gamma_init[l] = (num / den).clamp(min=0)

    return gamma_init

def MFOCUSS_initial_power(A, X):
    A2 = torch.mean(torch.sum(torch.abs(A)**2, dim=0))
    X2 = torch.mean(torch.sum(torch.abs(X)**2, dim=0))

    sMat = torch.abs(A.conj().T @ X)**2 / (A2 * X2)
    gamma_init = torch.mean(sMat, dim=1)

    return gamma_init, sMat

# ——————————————————————————————————————————————————————————————
# 5) Unfolded layer definition
# ——————————————————————————————————————————————————————————————
class LIAALayerConfig(nn.Module):
    def __init__(self, cfg):
        super().__init__()
        self.cfg = cfg
        L = cfg['L']

        if cfg.get('use_layernorm'):
            self.ln = nn.LayerNorm(L)
        if cfg.get('use_attention'):
            self.attn = nn.MultiheadAttention(embed_dim=1, num_heads=1)
        if cfg.get('use_gating'):
            self.gate = nn.Linear(L, L)
        if cfg['algorithm'] == 'IAA' and cfg.get('use_shrinkage'):
            self.lmbda = nn.Parameter(
                torch.tensor(cfg['lmbda_init'], dtype=torch.float32)
            )

        match cfg['algorithm']:
            case 'MFOCUSS':
                self._raw_p = nn.Parameter(torch.tensor(cfg['mfocuss_p_init'], dtype=torch.float32))
                self._raw_l = nn.Parameter(torch.tensor(cfg['mfocuss_lambda_init'], dtype=torch.float32))   
                self._delta = nn.Parameter(torch.tensor(cfg['delta_init'], dtype=torch.float32))                 
            case 'IAA':
                self._delta = nn.Parameter(torch.tensor(cfg['delta_init'], dtype=torch.float32))    
            case _:
                raise ValueError(f"Unknown algorithm: {cfg['algorithm']}")

          

    def _iaa_update(self, p, A, X):
        P = torch.diag(p).to(A.dtype).to(A.device)
        R = A @ P @ A.conj().T
        if self.cfg.get('use_ridge'):
            R += self.cfg['ridge_reg'] * torch.eye(
                R.size(0), dtype=A.dtype, device=A.device
            )
        Rinv = torch.linalg.inv(R)

        ARi = A.conj().T @ Rinv
        S = ARi @ X

        q = torch.sum(torch.abs(S)**2, dim=1)
        d = torch.real(torch.sum(ARi * A.conj().T, dim=1))
        v = q / (d + eps)

        if self.cfg.get('constrain_gamma'):
            δ = torch.sigmoid(self._delta)
        else:
            δ = self._delta
        
        if self.cfg.get('use_residual'):
            u = p + δ * (v - p)
        else:
            u = v
        
        if self.cfg.get('use_layernorm'):
            u_p = self.ln(u)
        else:
            u_p = u
        
        attn_out = torch.zeros_like(u_p)
        if self.cfg.get('use_attention'):
            seq, _ = self.attn(
                u_p.unsqueeze(1).unsqueeze(2),
                u_p.unsqueeze(1).unsqueeze(2),
                u_p.unsqueeze(1).unsqueeze(2),
            )
            attn_out = seq.squeeze()

        if self.cfg.get('use_gating'):
            g = torch.sigmoid(self.gate(u))
            r = g * v + (1 - g) * p
        else:
            r = u_p

        s = r + attn_out
        if self.cfg.get('use_shrinkage'):
            return F.relu(s - self.lmbda)
        else:
            return F.relu(s)

    def _mfocuss_update(self, gamma, A, X, unfolded_params: dict):

        raw_p = unfolded_params.get('raw_p')
        raw_l = unfolded_params.get('raw_lambda')

        if self.cfg.get('constrain_p_lam'):
            p_min = self.cfg['mfocuss_p_min']
            p_max = self.cfg['mfocuss_p_max']
            p = p_min + (p_max - p_min) * torch.sigmoid(raw_p)

            lam_min = self.cfg['mfocuss_lambda_min']
            lam_max = self.cfg['mfocuss_lambda_max']
            lam = lam_min + (lam_max - lam_min) * torch.sigmoid(raw_l)
        else:
            p   = raw_p 
            lam = raw_l

        # core update
        muVec = mfocuss_core(A, X, gamma, p, lam)
        
        gamma_raw = torch.sqrt(torch.sum(torch.abs(muVec)**2, dim=1))+eps

        # 3) (optional) residual mixing — here it's identity unless you introduce a learnable δ
        if self.cfg.get('constrain_gamma'):
            δ = torch.sigmoid(self._delta)
        else:
            δ = self._delta
        
        if self.cfg.get('use_residual'):
            u = gamma + δ * (gamma_raw - gamma)
        else:
            u = gamma_raw
  
        # 4) (optional) layer‐norm
        if self.cfg.get('use_layernorm', False):
            u = self.ln(u)

        # 5) (optional) attention
        attn_out = torch.zeros_like(u)
        if self.cfg.get('use_attention', False):
            seq, _ = self.attn(
                u.unsqueeze(1).unsqueeze(2),
                u.unsqueeze(1).unsqueeze(2),
                u.unsqueeze(1).unsqueeze(2),
            )
            attn_out = seq.squeeze()

        # 6) (optional) gating between new & old γ
        if self.cfg.get('use_gating', False):
            g = torch.sigmoid(self.gate(u))
            u = g * gamma_raw + (1 - g) * gamma

        # 7) combine with attention
        s = u + attn_out

        # 8) (optional) shrinkage thresholding by λ
        if self.cfg.get('use_shrinkage', False):
            gamma_new = F.relu(s - lam)
        else:
            gamma_new = F.relu(s)
        
        gamma_new = torch.nan_to_num(gamma_new, nan=0.0, posinf=1e6, neginf=0.0)
        gamma_new = torch.clamp(gamma_new, min=0.0)

        return gamma_new

    def forward(self, gamma, A, X, unfolded_params: dict):
        if self.cfg['algorithm'] == 'MFOCUSS':
            return self._mfocuss_update(gamma, A, X, unfolded_params)
        else:
            return self._iaa_update(gamma, A, X)

# ——————————————————————————————————————————————————————————————
# 5b) Classic iterative sanity‐check
# ——————————————————————————————————————————————————————————————
def classic_iaa(A, X, gamma_init, cfg, iters):
    """
    Classic IAA loop, but with renamed variables:
      • gamma_init : initial spectrum (was gamma_init)
      • gamma_vec  : current spectrum estimate
      • δ          : residual‐mix coefficient
    """
    # start from the provided initial spectrum
    gamma_vec = gamma_init.clone()

    # identity for ridge
    I = torch.eye(A.size(0), dtype=A.dtype, device=A.device)

    # mixing coefficient δ
    if cfg.get('constrain_delta', False):
        δ = torch.sigmoid(torch.tensor(cfg['delta_init'], device=A.device))
    else:
        δ = torch.tensor(cfg['delta_init'], device=A.device)

    for _ in range(iters):
        # 1) build weighted covariance
        P = torch.diag(gamma_vec).to(A.dtype).to(A.device)
        R = A @ P @ A.conj().T
        if cfg.get('use_ridge', True):
            R = R + cfg['ridge_reg'] * I

        # 2) invert & form v = q/d
        Rinv = torch.linalg.inv(R)
        ARi  = A.conj().T @ Rinv
        S    = ARi @ X
        q    = torch.sum(torch.abs(S)**2, dim=1)
        d    = torch.real(torch.sum(ARi * A.conj().T, dim=1))
        gamma_new    = q / (d + eps)

        # 3) residual mixing (if enabled)
        if cfg.get('use_residual', True):
            gamma_vec = gamma_vec + δ * (gamma_new - gamma_vec)
        else:
            gamma_vec = gamma_new

        # 4) enforce non-negativity
        gamma_vec = F.relu(gamma_vec)

    return gamma_vec

def normalize_cs_input(X_N, sf):
    mode = sf['nfMode']
    if mode == "Legacy":
        factor = sf['nfLegacyRMS'] / torch.sqrt(torch.sum(torch.abs(X_N)**2))
    elif mode == "Elements":
        factor = sf['nfRMS'] / torch.sqrt(torch.mean(torch.abs(X_N)**2, dim=1))
    else:
        factor = sf['nfRMS'] / torch.sqrt(torch.mean(torch.abs(X_N)**2))

    return X_N * factor, factor

def preserve_energy(X, powerSpec):
    E = torch.sum(torch.abs(X)**2)
    return powerSpec / (powerSpec.sum() + eps) * E

def mfocuss_core(A, Y, gamma, p, lam):
    I = torch.eye(A.size(0), dtype=A.dtype, device=A.device)

    W = torch.diag(gamma ** (1 - p/2)).to(A.dtype).to(A.device)
    AW = A @ W
    lam_eff = torch.clamp(torch.tensor(lam), min=1e-8)  # enforce strictly positive
    M  = AW @ AW.conj().T + lam_eff * I
    # M  = AW @ AW.conj().T + lam * I

    Q = AW.conj().T @ (torch.linalg.inv(M) @ Y)
    muVec = W @ Q

    return muVec

def Calc_frobenius_norm(muVec, muPrev):
    
    nom = torch.norm(muVec - muPrev, p='fro')
    den = torch.norm(muPrev, p='fro') + eps
    fro_norm = nom / den
    return fro_norm

def classic_mfocuss(A, X, cfg, max_iters):
    p_init = cfg['mfocuss_p_init']
    lambda_init = cfg['mfocuss_lambda_init']

    p_rate = cfg['p_decay_rate']
    lambda_rate = cfg['lambda_growth_rate']

    p_final = cfg['p_min_th']
    
    lambda_final = cfg['lambda_max_th']
    prune_gamma_db = cfg['prune_gamma_th_dB']
    epsilon = cfg['epsilon']
    s_norm_fact = cfg['s_norm_factor']

    Y, norm_factor = normalize_cs_input(X, s_norm_fact)

    _, muPrev = MFOCUSS_initial_power(A, X)

    p = p_init
    lam = lambda_init
    gamma_progress = []

    # Iterations Loop
    for it in range(1, max_iters + 1):
        gamma = torch.sqrt(torch.sum(torch.abs(muPrev)**2, dim=1))
        gamma_progress.append(gamma)

        # mfocuss_core
        muVec = mfocuss_core(A, Y, gamma, p, lam)

        # Calc_frobenius_norm
        fro_norm = Calc_frobenius_norm(muVec, muPrev)

        if fro_norm < epsilon:
            break

        muPrev = muVec

        # Update p and lambda per iteration
        p   = p_final + (p_init - p_final) * np.exp(-p_rate * it)
        lam = lambda_final + (lambda_init - lambda_final) * np.exp(-lambda_rate * it)

    gamma_out = torch.sqrt(torch.sum(torch.abs(muPrev)**2, dim=1))
    gamma_out = preserve_energy(X, gamma_out)

    db_gamma = 10 * torch.log10(gamma_out + eps)
    idx = db_gamma > prune_gamma_db
    gamma_out[db_gamma <= prune_gamma_db] = 10 ** (prune_gamma_db / 10)

    mu_pruned = muPrev[idx, :]

    s = mu_pruned / torch.sqrt(
        torch.sum(torch.abs(mu_pruned)**2, dim=1, keepdim=True) + eps)
    cov = torch.abs(s @ s.conj().T)

    return {
        'gamma':       gamma_out,
        'index':       idx,
        'mu':          mu_pruned,
        'cov':         cov,
        'm':           int(idx.sum().item()),
        'count':       it,
        'norm_factor': norm_factor,
    }

# ——————————————————————————————————————————————————————————————
# 6) Unfolded network definition
# ——————————————————————————————————————————————————————————————
class TransformerParamPredictor(nn.Module):
    """
    Transformer-based predictor for raw_p and raw_lambda from complex-valued input X.
    Input X is shaped (N, T_snap), interpreted as a sequence of T_snap vectors of dim N.
    """

    def __init__(self, cfg):
        super().__init__()
        self.cfg = cfg
        self.N = cfg['N']
        self.T_snap = cfg['T']
        self.num_layers = cfg['num_layers']
        self.embed_dim = 64
        self.num_heads = 4
        self.ff_dim = 128

        # Project real+imag components of X into embedding
        self.input_proj = nn.Linear(2 * self.N, self.embed_dim)

        # Transformer encoder
        encoder_layer = nn.TransformerEncoderLayer(
            d_model=self.embed_dim,
            nhead=self.num_heads,
            dim_feedforward=self.ff_dim,
            batch_first=True,
        )
        self.encoder = nn.TransformerEncoder(encoder_layer, num_layers=2)

        # Final projection to raw_p and raw_lambda
        self.out_proj = nn.Linear(self.embed_dim, 2 * self.num_layers)

    def forward(self, X):
        # X: (N, T_snap), complex
        x_real = X.real.transpose(0, 1)  # (T_snap, N)
        x_imag = X.imag.transpose(0, 1)  # (T_snap, N)
        x = torch.cat([x_real, x_imag], dim=-1)  # (T_snap, 2N)

        # Add batch dim: (1, T_snap, 2N)
        x = self.input_proj(x.unsqueeze(0))  # (1, T_snap, embed_dim)
        x = self.encoder(x)  # (1, T_snap, embed_dim)

        # Mean pooling across time dimension
        x_pooled = x.mean(dim=1)  # (1, embed_dim)

        # Output projection
        out = self.out_proj(x_pooled).squeeze(0)  # (2 * T,)

        unfolded_params = {
            'raw_p': out[:self.num_layers],
            'raw_lambda': out[self.num_layers:]
        }
        return unfolded_params

class MixerParamPredictor(nn.Module):
    """
    MLP-Mixer-inspired model for parameter prediction from input X (N × T_snap).
    """

    def __init__(self, cfg):
        super().__init__()
        self.cfg = cfg
        self.N = cfg['N']
        self.T_snap = cfg['T']
        self.num_layers = cfg['num_layers']
        hidden_dim = 64
        token_dim = 128
        channel_dim = 128

        self.norm = nn.LayerNorm(2 * self.T_snap)

        # Project real+imag to tokens
        self.token_mlp = nn.Sequential(
            nn.Linear(2 * self.T_snap, token_dim),
            nn.GELU(),
            nn.Linear(token_dim, 2 * self.T_snap)
        )

        self.channel_mlp = nn.Sequential(
            nn.Linear(self.N, channel_dim),
            nn.GELU(),
            nn.Linear(channel_dim, self.N)
        )

        self.out_proj = nn.Sequential(
            nn.Flatten(),
            nn.Linear(self.N * 2 * self.T_snap, 128),
            nn.ReLU(),
            nn.Linear(128, 2 * self.num_layers)
        )

    def forward(self, X):
        # X: (N, T_snap), complex
        x_real = X.real  # (N, T)
        x_imag = X.imag  # (N, T)
        x = torch.cat([x_real, x_imag], dim=-1)  # (N, 2T)

        # Token mixing
        x = self.token_mlp(self.norm(x)) + x

        # Channel mixing
        x = x.transpose(0, 1)  # (2T, N)
        x = self.channel_mlp(x) + x
        x = x.transpose(0, 1)  # (N, 2T)

        out = self.out_proj(x)  # (2 * T,)
        unfolded_params = {
            'raw_p': out[:self.num_layers],
            'raw_lambda': out[self.num_layers:]
        }
        return unfolded_params

class MLPParamPredictor(nn.Module):
    def __init__(self, cfg):
        super().__init__()
        self.cfg = cfg
        self.N = cfg['N']
        self.T_snap = cfg['T']
        self.num_layers = cfg['num_layers']

        input_dim = 2 * self.N * self.T_snap
        hidden_dim = 128
        output_dim = 2 * self.num_layers

        self.model = nn.Sequential(
            nn.Linear(input_dim, hidden_dim),
            nn.ReLU(),
            nn.Linear(hidden_dim, hidden_dim),
            nn.ReLU(),
            nn.Linear(hidden_dim, output_dim)
        )

    def forward(self, X):
        x_real = X.real.flatten()
        x_imag = X.imag.flatten()
        x = torch.cat([x_real, x_imag], dim=0).unsqueeze(0)
        out = self.model(x).squeeze(0)

        return {
            'raw_p': out[:self.cfg['num_layers']],
            'raw_lambda': out[self.cfg['num_layers']:]
        }

class ComplexConvParamPredictor(nn.Module):
    """
    CNN-based model that processes complex input using real+imag channel stacking.
    Input: X ∈ ℂ^{5×8} → stacked to (2, 5, 8)
    Output: raw_p and raw_lambda for each unfolded layer.
    """

    def __init__(self, cfg):
        super().__init__()
        self.cfg = cfg
        self.N = cfg['N']
        self.T_snap = cfg['T']
        self.num_layers = cfg['num_layers']

        self.cnn = nn.Sequential(
            nn.Conv2d(in_channels=2, out_channels=16, kernel_size=(3, 3), padding=1),
            nn.ReLU(),
            nn.Conv2d(16, 32, kernel_size=(3, 3), padding=1),
            nn.ReLU(),
            nn.AdaptiveAvgPool2d((2, 2)),  # reduce spatial size
        )

        self.mlp = nn.Sequential(
            nn.Flatten(),
            nn.Linear(32 * 2 * 2, 128),
            nn.ReLU(),
            nn.Linear(128, 2 * self.num_layers)
        )

    def forward(self, X):
        """
        Input:
            X: complex-valued input (5, 8)
        Output:
            dict with 'raw_p' and 'raw_lambda', each of shape (num_layers,)
        """
        x_real = X.real.unsqueeze(0)  # (1, 5, 8)
        x_imag = X.imag.unsqueeze(0)  # (1, 5, 8)
        x = torch.cat([x_real, x_imag], dim=0).unsqueeze(0)  # (1, 2, 5, 8)

        features = self.cnn(x)  # (1, 32, 2, 2)
        out = self.mlp(features).squeeze(0)  # (2*T,)

        return {
            'raw_p':      out[:self.num_layers],
            'raw_lambda': out[self.num_layers:]
        }
            
class FullyConnectedParamPredictor(nn.Module):
    """
    Fully connected neural network to predict raw_p and raw_lambda values
    for each of the T unfolded iterations, given the 5×8 phasor input.
    """

    def __init__(self, cfg):
        super().__init__()
        self.cfg = cfg
        self.N = cfg['N']  # number of antenna elements
        self.T_snap = cfg['T']  # number of snapshots
        self.num_layers = cfg['num_layers']  # T: unfolding depth

        input_dim = 2 * self.N * self.T_snap  # Real + Imag parts
        hidden_dim = 128  # can be tuned
        output_dim = 2 * self.num_layers  # 2 scalars per layer: raw_p, raw_lambda

        self.model = nn.Sequential(
            nn.Linear(input_dim, hidden_dim),
            nn.ReLU(),
            nn.Linear(hidden_dim, hidden_dim),
            nn.ReLU(),
            nn.Linear(hidden_dim, output_dim)
        )

    def forward(self, X):
        """
        Input:
            X: torch.Tensor of shape (N, T_snap), dtype=torch.complex64
        Output:
            raw_params: torch.Tensor of shape (2, num_layers)
        """
        # Flatten and split real/imag
        x_real = X.real.flatten()
        x_imag = X.imag.flatten()
        x = torch.cat([x_real, x_imag], dim=0).unsqueeze(0)  # shape (1, 2*N*T)

        out = self.model(x)  # shape (1, 2*T)
        raw_params = out.view(2, self.num_layers)  # shape (2, T)
        return raw_params


class LIAA_Net_Config(nn.Module):
    def __init__(self, cfg):
        super().__init__()
        self.cfg = cfg
        self.num_layers = cfg['num_layers']
        self.param_predictor_module = cfg['param_predictor_module']
        self.layers = nn.ModuleList([LIAALayerConfig(cfg) for _ in range(self.num_layers)])
        self.param_predictor = self.build_param_predictor(cfg)

    def forward(self, gamma_init, A, X):
        gamma = gamma_init
        gamma_progress = []
        # if config['debug_plot']:
        #     Plot_gamma(pow2db(gamma_init))

        if self.cfg.get('preserve_energy'):
            gamma = preserve_energy(X, gamma)

        # Norm Factor
        # s_norm_fact = self.cfg['s_norm_factor']
        # X, _ = normalize_cs_input(X, s_norm_fact)
        # Show model summary
        is_complex = True  # since X is complex-valued
        show_model_summary(self.param_predictor, X.shape, model_name=config['param_predictor'].upper(), is_complex=is_complex)


        # Predict unfolding parameters from the input signal X
        raw_params = self.param_predictor(X)  # shape (2, num_layers)
        raw_p_seq = raw_params[0]  # shape (T,)
        raw_l_seq = raw_params[1]  # shape (T,)

        for t, layer in enumerate(self.layers):
            unfolded_params = {
                'raw_p': raw_p_seq[t],
                'raw_lambda': raw_l_seq[t],
            }
            gamma = layer(gamma, A, X, unfolded_params)
            gamma_progress.append(gamma)

        return gamma, gamma_progress

    def build_param_predictor(self, cfg):
        match self.param_predictor_module:
            case 'FullyConnected': 
                return FullyConnectedParamPredictor(cfg)
            case 'MLP':
                return MLPParamPredictor(cfg)
            case 'Transformer':
                return TransformerParamPredictor(cfg)
            case 'Mixer':
                return MixerParamPredictor(cfg)
            case 'ComplexConv':
                return ComplexConvParamPredictor(cfg)
            case _:
                raise ValueError(f"Unknown param_predictor: {cfg['param_predictor']}")
            
# ——————————————————————————————————————————————————————————————
# 7) Loss & peak‐finder
# ——————————————————————————————————————————————————————————————
def nmse_loss(g_hat, g_true, eps=1e-12):
    num = torch.sum((g_hat - g_true)**2, dim=1)
    den = torch.sum(g_true**2, dim=1) + eps
    return torch.mean(num / den)

def conv_nmse_loss(g_est: torch.Tensor,
                   g_true: torch.Tensor,
                   cfg: dict) -> torch.Tensor:
    """
    1) Build a triangular kernel of support [-th_md_deg, +th_md_deg]
       in the same binning as theta_grid.
    2) Convolve g_est and g_true with that kernel (padding so output
       stays length L).
    3) Normalize each by its own peak absolute value.
    4) Compute NMSE( g_est_norm, g_true_norm ).
    """
    theta_grid = np.asarray(cfg['theta_grid_deg'], dtype=float)
    th_md_deg  = cfg['th_md_deg']

    # 1) compute grid spacing in degrees and kernel half‐width in bins
    delta = float(theta_grid[1] - theta_grid[0])
    r     = int(round(th_md_deg / delta))

    # 2) if threshold is < 1 bin, just normalize & do plain NMSE
    if r >= 1:
        # 3) build triangular convolution kernel
        device = g_est.device
        kernel = 1.0 - torch.abs(torch.arange(-r, r+1, device=device, dtype=torch.float32))/r
        kernel = kernel.clamp(min=0.0)
        kernel = kernel / (kernel.sum() + eps)
        kernel = kernel.view(1,1,-1)  # for conv1d

        # 4) convolve
        g_est  = F.conv1d(g_est .view(1,1,-1), kernel, padding=r).view(-1)
        g_true = F.conv1d(g_true.view(1,1,-1), kernel, padding=r).view(-1)

  
    # 5) normalize each by its own max‐abs
    max_est  = g_est .abs().max()
    max_true = g_true.abs().max()
    g_est_n  = g_est  / (max_est  + eps)
    g_true_n = g_true / (max_true + eps)

    # Calculate Loss
    loss = nmse_loss(
        g_est_n.unsqueeze(0),
        g_true_n.unsqueeze(0),
        eps)

    # 6) optional debug plot
    if cfg.get('debug_loss', False):
        # you may want to pass your real results/epoch/loss into this call
        Plot_gamma_hat_vs_true(
            g_true_n,   g_est_n,
            results=None,
            epoch=None,
            loss=loss
        )

    # 7) final NMSE
    return loss

def wrap_to_180(x):
    return ((x + 180) % 360) - 180

def keep_only_peaks(g_hat: torch.Tensor, peaks_inds) -> torch.Tensor:
    """
    Return a new tensor of the same shape as `g_hat`, but zero everywhere
    except at the indices in `peaks_inds`, which retain their original values.
    """
    # make a zero‐tensor
    new_g = torch.zeros_like(g_hat)

    # turn peaks_inds into a LongTensor on the same device
    if isinstance(peaks_inds, np.ndarray):
        idx = torch.from_numpy(peaks_inds).long().to(g_hat.device)
    else:
        idx = torch.tensor(peaks_inds, dtype=torch.long, device=g_hat.device)

    # copy over only the peak values
    new_g[idx] = g_hat[idx]

    return new_g

def calc_doa_results(g_hat: torch.Tensor,
                     cfg: dict) -> dict:
    """
    1) Pick the top-K peaks from g_hat
    2) Build the G×E error matrix
    3) Greedily match the smallest error to assign each estimate at most once
    4) Compute MD, FA, MAE, and record per-GT assigned estimate
    """
    # unpack
    gt = np.asarray(cfg['gt_angles_deg'], dtype=float)   # (G,)
    th = cfg['th_md_deg']
    max_est_peaks  = cfg['max_est_peaks']

    # 1) find your K largest peaks
    idxs, est = find_peaks(g_hat, max_est_peaks)                   # est_t is torch.Tensor(E,)

    # get a ĝ that is zero except at your detected DOA bins
    g_est = keep_only_peaks(g_hat, idxs)

    G = len(gt)
    E = len(est)
    # 2) compute G×E wrapped-difference matrix
    #    err[i,j] = | wrap(gt[i] - est[j]) |
    err = np.abs(wrap_to_180(gt[:,None] - est[None,:]))  # shape (G,E)

    # 3) prepare our storage
    dtctDOA         = np.full(G, np.nan)
    asignedRel      = np.full(G, np.nan)
    dtctMask        = np.zeros(G, dtype=bool)
    M               = err.copy()

    # greedy assignment loop
    while True:
        if np.all(np.isnan(M)):
            break
        minVal = np.nanmin(M)
        if minVal > th:
            break
        # where is that minimum?
        flat   = np.nanargmin(M)
        i_gt, i_est = np.unravel_index(flat, M.shape)
        # assign
        dtctMask[i_gt]      = True
        dtctDOA[i_gt]       = est[i_est]
        asignedRel[i_gt]    = gt[i_gt]
        # mask out that row & column
        M[i_gt, :]  = np.nan
        M[:, i_est] = np.nan

    # 4) now compute your metrics
    dtctCnt = dtctMask.sum()
    md_rate = 1 - dtctCnt / G
    fa_rate = (E - dtctCnt) / E if E>0 else np.nan
    dtctErr = dtctDOA - asignedRel
    absErr  = np.abs(dtctErr)
    mae_deg = np.nanmean(absErr) if dtctCnt>0 else np.nan

    results = {
        'GT_DOA_deg'       : gt,            # shape (G,)
        'Dtct_DOA_deg'     : dtctDOA,       # shape (G,)
        'Est_DOA_deg'      : est,           # shape (E,)
        'Peaks_inds'       : idxs,          # shape (G,)
        'Abs_Err_deg'      : absErr,        # shape (G,)
        'MAE_deg'          : mae_deg,
        'MD'               : md_rate,
        'FA'               : fa_rate
    }
    
    return results, g_est

# ——————————————————————————————————————————————————————————————
#  Find Peaks
# ——————————————————————————————————————————————————————————————
def find_peaks_local(gamma: np.ndarray,
                     theta_grid: np.ndarray,
                     K: int,
                     min_height: float = None,
                     min_distance: int = 1):
    """
    1) Finds local maxima in `gamma` (using scipy)
    2) Optionally thresholds them by `min_height`
    3) Enforces a minimum index separation `min_distance`
    4) Returns the top‐K by amplitude
    """
    # 1) find *all* local peak indices
    peaks, props = sp_find_peaks(gamma,
                                 height=min_height,
                                 distance=min_distance)

    if peaks.size == 0:
        return np.array([], dtype=int), np.array([])

    # 2) sort those peaks by their gamma value
    peak_vals = gamma[peaks]
    order     = np.argsort(peak_vals)[::-1]

    # 3) take the top‐K of them
    order     = order[:K]
    chosen    = peaks[order]

    # 4) return both indices and the angle locations
    return chosen, theta_grid[chosen]

def Calc_min_height(gamma_np):
    gap_db = config.get('gap_from_peak_dB', None)
    if gap_db is not None:
        gamma_db       = 10 * np.log10(gamma_np + eps)
        max_db         = np.nanmax(gamma_db)
        min_height_db  = max_db - gap_db
        min_height     = 10 ** (min_height_db / 10)
    else:
        min_height = None

    return min_height

def find_peaks(g_hat, K):
    gamma_np   = g_hat.detach().cpu().numpy()
    theta_grid = config['theta_grid_deg']
    deg_sep    = config['min_df_dist_combine_deg']

    # 1) compute your grid resolution (assumes uniform spacing)
    Δθ = np.mean(np.diff(theta_grid))

    # 2) convert the angular separation into an integer index distance
    min_peak_distance = max(1, int(np.round(deg_sep / Δθ)))
    
    # 2) compute minimum height (in dB) = max_peak - gap_from_peak_dB
    min_height = Calc_min_height(gamma_np)

    # 3) call your local‐peak finder with that index separation
    idxs, est_angles = find_peaks_local(
        gamma_np,
        theta_grid,
        K,
        min_height=min_height,
        min_distance=min_peak_distance
    )
    return idxs, est_angles

# ——————————————————————————————————————————————————————————————
# A) wrapTo180
# ——————————————————————————————————————————————————————————————
def wrap_to_180(angle):
    """
    Wrap angle(s) into (–180, 180] degrees.
    Works element‐wise on arrays.
    """
    return ((angle + 180) % 360) - 180

# ——————————————————————————————————————————————————————————————
# 8) Synthetic DOA Dataset
# ——————————————————————————————————————————————————————————————
class DOADataset(Dataset):
    def __init__(self, num_samples, cfg):
        super().__init__()
        self.num_samples = num_samples
        self.cfg = cfg
        self.K = cfg['K']
        self.theta_rad = torch.tensor(
            cfg['theta_grid_rad'], dtype=torch.float32
        )
        self.theta_deg = cfg['theta_grid_deg']
        self.N, self.L, self.T = cfg['N'], cfg['L'], cfg['T']
        snr_min, snr_max = self.cfg['SNR_dB_range']
        snr_db           = np.random.uniform(snr_min, snr_max)
        snr_lin          = 10 ** (snr_db / 10)
        # snr_lin = 10**(cfg['SNR_dB']/10)
        self.sigma = np.sqrt(self.K / snr_lin)

        self.A = torch.exp(
            -1j * 2 * np.pi * 0.5 *
            torch.outer(
                torch.arange(self.N, dtype=torch.float32),
                torch.sin(self.theta_rad)
            )
        )

    def __len__(self):
        return self.num_samples

    def __getitem__(self, idx):
        if self.cfg.get('gt_angles_deg') is not None:
            thetas_true = self.cfg['gt_angles_deg'].clone().detach() * np.pi/180.0
            # thetas_true = torch.tensor(
            #     self.cfg['gt_angles_deg'],
            #     dtype=torch.float32
            # ) * np.pi/180.0
        else:
            self.cfg['gt_angles_deg'] = torch.rand(self.K) * 120 - 60
            thetas_true = self.cfg['gt_angles_deg'] * np.pi/180.0

        A_true = torch.exp(
            -1j * 2 * np.pi * 0.5 *
            torch.outer(
                torch.arange(self.N, dtype=torch.float32),
                torch.sin(thetas_true)
            )
        )

        S = (torch.randn(self.K, self.T) +
             1j * torch.randn(self.K, self.T)) / np.sqrt(2)
        noise = self.sigma * (
            torch.randn(self.N, self.T) +
            1j * torch.randn(self.N, self.T)
        ) / np.sqrt(2)
        X = A_true @ S + noise

        power = torch.sqrt(torch.mean(torch.abs(S)**2, dim=1))

        gamma_true = torch.zeros(self.L)
        for j, th  in enumerate(thetas_true):
            i = torch.argmin(torch.abs(self.theta_rad - th))
            
            gamma_true[i] = power[j]

        match self.cfg['algorithm']:
            case 'MFOCUSS':
                gamma_init, _ = MFOCUSS_initial_power(self.A, X)
            case 'IAA':
                gamma_init = IAA_initial_power(self.A, X)

        return self.A, X, gamma_init, gamma_true

# ——————————————————————————————————————————————————————————————
# 9) Train
# ——————————————————————————————————————————————————————————————
def train_model(net, loader, optimizer, cfg):
    net.train()

    # accumulate per‐sample histories here
    history = {}   # name -> list of lists
    tags    = []   # e.g. ["e1_b1","e1_b2",…]
    log_dir = "C:/Data/Unfolding/Training_logs"

    for epoch in range(1, cfg['num_epochs']+1):
        for batch_idx, batch in enumerate(tqdm(loader, desc=f"Epoch {epoch}"), start=1):
            A, X, g0, g_true = batch
            A      = A[0]; 
            X      = X[0]
            g0     = g0[0]; 
            g_true = g_true[0]

            # Zero Gradients Optimizer
            optimizer.zero_grad()

            # 1) Apply Network
            g_hat, gamma_progress = net(g0, A, X)

            if cfg['debug_plot']:
                Plot_gamma(pow2db(g_hat))
                
            # 2) Compute & gather DOA results for this sample
            results, g_est = calc_doa_results(g_hat, cfg)

            if cfg['debug_loss']:
                Plot_gamma_hat_vs_true(g0, g_hat, results)

            # Show_power_spectrum
            if cfg['debug_plot']:
                Show_power_spectrum(gamma_progress, results, g_true, g_est)
                
            # 3) Convolutional NMSE Loss
            loss = conv_nmse_loss(g_est, g_true, cfg)

            # 4) Back-propogation
            loss.backward()

            # 5) clip_grad_norm_
            torch.nn.utils.clip_grad_norm_(net.parameters(), max_norm=1.0)

            # Optimizer Step
            optimizer.step()

            # 6) gather **per‐layer** learnable params exactly as before
            Update_history(history, results, net)

            # 7) tag this column
            tags.append(f"e{epoch}_b{batch_idx}")

    # once all epochs & batches done, dump to Excel:
    save_history(history, tags, log_dir, "Training_results.xlsx", "Train")

# ——————————————————————————————————————————————————————————————
# 10) Test
# ——————————————————————————————————————————————————————————————
def test_model(net: nn.Module, cfg: dict):
    net.eval()

    # how many test samples?
    num_test = cfg.get('num_test', cfg['num_train']//5)

    # prepare your dataset and loader (batch_size=1 so we can tag each sample)
    ds     = DOADataset(num_test, cfg)
    loader = DataLoader(ds, batch_size=1, shuffle=False)

    # history will collect one entry per sample
    history = {}
    tags    = []
    log_dir = "C:/Data/Unfolding/Test_logs"

    for sample_idx, batch in enumerate(loader, start=1):
        A, X, g_init, g_true = batch
        A      = A[0]
        X      = X[0]
        g_init = g_init[0]
        g_true = g_true[0]

        with torch.no_grad():
            g_hat, gamma_progress = net(g_init, A, X)

        # compute exactly the same per-sample results dict
        results, g_est = calc_doa_results(g_hat, cfg)
        
        if cfg['debug_plot']:
            Show_power_spectrum(gamma_progress, results, g_true, g_est)

        # tag this sample so columns read sample_1, sample_2, ...
        tags.append(f"sample_{sample_idx}")

        # flatten & append each field into history
        for name, arr in results.items():
            a = arr
            if hasattr(a, "detach"):
                a = a.detach().cpu().numpy()
            else:
                a = np.asarray(a)
            history.setdefault(name, []).append(a.ravel().tolist())

    # at the end, write them all out to Test_logs/test_results.xlsx
    save_history(history, tags, log_dir, "Test_results.xlsx", "Test")

# ——————————————————————————————————————————————————————————————
# 11) Main Classic
# ——————————————————————————————————————————————————————————————
def classic():
    ds = DOADataset(1, config)
    A, X, g_init, g_true = ds[0]

    match config['algorithm']:
        case 'MFOCUSS':
            out = classic_mfocuss(A, X, config, config['classic_iters'])
            g_hat = out['gamma']
        case 'IAA':
            g_hat = classic_iaa(A, X, g_init, config, config['classic_iters'])

    # calc_doa_results
    results, g_est = calc_doa_results(g_hat, config)

    # Plot_gamma_hat_vs_true
    Plot_gamma_hat_vs_true(g_true, g_est, results)

# ——————————————————————————————————————————————————————————————
# 12) Main Unfolded
# ——————————————————————————————————————————————————————————————
def unfolded():
    ds = DOADataset(config['num_train'], config)

    # Load Data
    loader = DataLoader(ds, batch_size=config['batch_size'], shuffle=True)

    # Create Network
    net = LIAA_Net_Config(config)

    # Op[timize
    opt = torch.optim.Adam(
        net.parameters(),
        lr=config['learning_rate'],
        weight_decay=config['weight_decay']
    )
    
    # Train
    train_model(net, loader, opt, config)

    # Test
    test_model(net, config)

# ——————————————————————————————————————————————————————————————
# 13) Main entrypoint
# ——————————————————————————————————————————————————————————————
def main():
    Update_gt_angles(config)

    match config['mode']:
        case 'classic':
            classic()
        case 'unfolded':
            unfolded()

# ——————————————————————————————————————————————————————————————
# 13) Call Main
# ——————————————————————————————————————————————————————————————
if __name__ == "__main__":
    main()
